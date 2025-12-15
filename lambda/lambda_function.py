"""
ChargeScrub Lambda Function v1.3.2
Processes hospital billing charge spreadsheets for duplicate/overlapping E/M charges.
PHI-safe: No logging of file contents.
"""
import json
import base64
import io
import os
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# CORS origin - set via Lambda environment variable, or defaults to * for development
# In production, set ALLOWED_ORIGIN env var to your domain (e.g., https://scrub.yourdomain.com)
ALLOWED_ORIGIN = os.environ.get('ALLOWED_ORIGIN', '*')

# Code categories
INITIAL_CODES = ['99221', '99222', '99223']
SUBSEQUENT_CODES = ['99231', '99232', '99233']
DISCHARGE_CODES = ['99238', '99239']
CRITICAL_CODES = ['99291']
ACP_CODES = ['99497']

def assign_recommendations(group):
    """Apply billing rules to a patient-DOS group."""
    codes = set(group['PROCEDURECODE'])
    has_initial = any(c in INITIAL_CODES for c in codes)
    has_subsequent = any(c in SUBSEQUENT_CODES for c in codes)
    has_discharge = any(c in DISCHARGE_CODES for c in codes)
    has_critical = any(c in CRITICAL_CODES for c in codes)
    has_acp = any(c in ACP_CODES for c in codes)
    combo = '+'.join(sorted(codes)) if len(codes) > 1 else group['PROCEDURECODE'].iloc[0]
    
    # Get the provider who did the initial (for ACP rule)
    initial_provider = None
    if has_initial:
        initial_rows = group[group['PROCEDURECODE'].isin(INITIAL_CODES)]
        if len(initial_rows) > 0:
            initial_provider = initial_rows['BILLINGPROVIDER'].iloc[0]
    
    recs = []
    for idx, row in group.iterrows():
        code = row['PROCEDURECODE']
        provider = row.get('BILLINGPROVIDER', None)
        rec = "Charge"
        
        # Rule 1: Always deny 99499
        if code == '99499':
            rec = "Deny"
            recs.append(rec)
            continue
        
        # Rule 2: ACP (99497) - Charge unless same provider did the admit
        if code == '99497':
            if has_initial and initial_provider and provider == initial_provider:
                rec = "Deny"
            else:
                rec = "Charge"
            recs.append(rec)
            continue
        
        # Rule 3: Critical + Subsequent + Discharge (no initial) → Critical charges, others deny
        # Patient was to be discharged but became critical - critical supersedes
        if has_critical and has_subsequent and has_discharge and not has_initial:
            if code in CRITICAL_CODES:
                rec = "Charge"
            elif code in SUBSEQUENT_CODES or code in DISCHARGE_CODES:
                rec = "Deny"
        
        # Rule 4: Critical + Subsequent only (no discharge, no initial) → Review both (timeline unknown)
        elif has_critical and has_subsequent and not has_discharge and not has_initial:
            if code in CRITICAL_CODES or code in SUBSEQUENT_CODES:
                rec = "Review"
        
        # Rule 5: Initial + Critical both present → Review both (timeline unknown)
        elif has_initial and has_critical:
            if code in INITIAL_CODES or code in CRITICAL_CODES:
                rec = "Review"
            elif code in SUBSEQUENT_CODES or code in DISCHARGE_CODES:
                rec = "Deny"
        
        # Rule 6: Initial present, NO critical → Charge initial, Deny subsequent/discharge
        elif has_initial and not has_critical:
            if code in INITIAL_CODES:
                rec = "Charge"
            elif code in SUBSEQUENT_CODES or code in DISCHARGE_CODES:
                rec = "Deny"
        
        # Rule 7: Critical only (no initial, no subsequent) → Charge critical
        elif has_critical and not has_initial and not has_subsequent:
            if code in CRITICAL_CODES:
                rec = "Charge"
            elif code in DISCHARGE_CODES:
                rec = "Deny"
        
        # Rule 8: Discharge present, no initial, no critical → Charge discharge, Deny subsequent
        elif has_discharge and not has_initial and not has_critical:
            if code in DISCHARGE_CODES:
                rec = "Charge"
            elif code in SUBSEQUENT_CODES:
                rec = "Deny"
        
        recs.append(rec)
    
    group['Recommendation'] = recs
    group['Code Combo'] = combo
    return group

def process_excel(file_bytes, sheet_name=None):
    """Process an Excel file and return the scrubbed version."""
    # Read into pandas
    if sheet_name:
        df = pd.read_excel(io.BytesIO(file_bytes), sheet_name=sheet_name)
    else:
        df = pd.read_excel(io.BytesIO(file_bytes), sheet_name=0)
    
    # Clean columns
    df.columns = df.columns.str.strip()
    df['PROCEDURECODE'] = df['PROCEDURECODE'].astype(str).str.strip()
    df['DOS'] = pd.to_datetime(df['DOS'])
    df['PATIENT NAME'] = df['PATIENT NAME'].str.upper().str.strip()
    
    # Apply recommendations - reset index to preserve grouping columns
    df = df.groupby(['PATIENT NAME', 'DOS'], group_keys=False).apply(assign_recommendations)
    
    # Fix Code Combo for singles
    single_mask = df.groupby(['PATIENT NAME', 'DOS'])['PROCEDURECODE'].transform('count') == 1
    df.loc[single_mask, 'Code Combo'] = df.loc[single_mask, 'PROCEDURECODE']
    
    # Add notes for cases that need timeline review
    df['Notes'] = ''
    for idx, row in df.iterrows():
        combo_codes = set(row['Code Combo'].split('+'))
        has_initial_in_combo = any(c in INITIAL_CODES for c in combo_codes)
        has_critical_in_combo = any(c in CRITICAL_CODES for c in combo_codes)
        has_subsequent_in_combo = any(c in SUBSEQUENT_CODES for c in combo_codes)
        
        if row['Recommendation'] == 'Review':
            # Initial + Critical combo
            if has_initial_in_combo and has_critical_in_combo:
                df.at[idx, 'Notes'] = 'Check timeline: If Critical hours after Initial = both bill; otherwise may be redundant'
            # Critical + Subsequent combo (no initial)
            elif has_critical_in_combo and has_subsequent_in_combo and not has_initial_in_combo:
                df.at[idx, 'Notes'] = 'Check timeline: Subsequent AM → Critical PM = both bill; Critical AM → Subsequent PM = only Critical'
    
    # Create styled Excel output
    wb = Workbook()
    ws = wb.active
    ws.title = "Scrubbed Charges"
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1e40af")
    charge_fill = PatternFill("solid", fgColor="dcfce7")
    deny_fill = PatternFill("solid", fgColor="fee2e2")
    review_fill = PatternFill("solid", fgColor="dbeafe")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Write headers
    headers = list(df.columns)
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    # Write data rows
    for row_idx, row in enumerate(df.values, 2):
        for col_idx, value in enumerate(row, 1):
            # Handle datetime
            if isinstance(value, pd.Timestamp):
                value = value.strftime('%Y-%m-%d')
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            
            # Color code recommendation column
            if headers[col_idx - 1] == 'Recommendation':
                if value == 'Charge':
                    cell.fill = charge_fill
                elif value == 'Deny':
                    cell.fill = deny_fill
                elif value == 'Review':
                    cell.fill = review_fill
    
    # Auto-fit columns
    for col_idx, header in enumerate(headers, 1):
        max_len = len(str(header))
        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 2, 50)
    
    # Add summary sheet
    ws_summary = wb.create_sheet("Summary")
    
    # Calculate stats
    total_charges = len(df)
    charge_count = len(df[df['Recommendation'] == 'Charge'])
    deny_count = len(df[df['Recommendation'] == 'Deny'])
    review_count = len(df[df['Recommendation'] == 'Review'])
    unique_patients = df['PATIENT NAME'].nunique()
    
    summary_data = [
        ["ChargeScrub Analysis Summary", ""],
        ["Generated", datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
        ["", ""],
        ["RESULTS", ""],
        ["Total Charges", total_charges],
        ["Unique Patients", unique_patients],
        ["", ""],
        ["RECOMMENDATIONS", ""],
        ["Charge", charge_count],
        ["Deny", deny_count],
        ["Review", review_count],
        ["", ""],
        ["ALGORITHM LOGIC (v1.3)", ""],
        ["", ""],
        ["Rule 1: 99499 (Unlisted E/M)", "Always DENY"],
        ["Rule 2: 99497 (ACP)", "CHARGE unless same provider did admit → DENY"],
        ["", ""],
        ["Rule 3: Critical + Subsequent + Discharge (no Initial)", ""],
        ["   - Critical (99291)", "CHARGE"],
        ["   - Subsequent (99231-99233)", "DENY"],
        ["   - Discharge (99238-99239)", "DENY"],
        ["   Note: Patient was to discharge but became critical", ""],
        ["", ""],
        ["Rule 4: Critical + Subsequent Only (no Discharge/Initial)", ""],
        ["   - Critical (99291)", "REVIEW"],
        ["   - Subsequent (99231-99233)", "REVIEW"],
        ["   Note: Timeline determines billing", ""],
        ["", ""],
        ["Rule 5: Initial + Critical Both Present", ""],
        ["   - Initial (99221-99223)", "REVIEW"],
        ["   - Critical (99291)", "REVIEW"],
        ["   - Subsequent", "DENY"],
        ["   - Discharge", "DENY"],
        ["   Note: Timeline determines billing", ""],
        ["", ""],
        ["Rule 6: Initial Present, No Critical", ""],
        ["   - Initial code", "CHARGE"],
        ["   - Subsequent", "DENY"],
        ["   - Discharge", "DENY"],
        ["", ""],
        ["Rule 7: Critical Only (no Initial/Subsequent)", ""],
        ["   - Critical (99291)", "CHARGE"],
        ["   - Discharge", "DENY"],
        ["", ""],
        ["Rule 8: Discharge Only (no Initial/Critical)", ""],
        ["   - Discharge code", "CHARGE"],
        ["   - Subsequent", "DENY"],
        ["", ""],
        ["Rule 9: Single Codes", "CHARGE"],
    ]
    
    for row_idx, row_data in enumerate(summary_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_summary.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1:
                cell.font = Font(bold=True, size=14)
            elif row_idx in [7, 12]:
                cell.font = Font(bold=True)
    
    ws_summary.column_dimensions['A'].width = 50
    ws_summary.column_dimensions['B'].width = 25
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue(), {
        'total': total_charges,
        'charge': charge_count,
        'deny': deny_count,
        'review': review_count,
        'patients': unique_patients
    }

def parse_multipart(body, boundary):
    """Parse multipart form data and extract file bytes."""
    # Split by boundary
    parts = body.split(f'--{boundary}'.encode())
    
    for part in parts:
        # Skip empty parts and closing boundary
        if not part or part.strip() == b'' or part.strip() == b'--':
            continue
        
        # Check if this part contains a file
        if b'filename=' in part:
            # Find the blank line that separates headers from content
            header_end = part.find(b'\r\n\r\n')
            if header_end == -1:
                header_end = part.find(b'\n\n')
                content_start = header_end + 2
            else:
                content_start = header_end + 4
            
            if header_end != -1:
                file_bytes = part[content_start:]
                # Remove trailing CRLF and boundary markers
                while file_bytes.endswith(b'\r\n') or file_bytes.endswith(b'\n'):
                    file_bytes = file_bytes.rstrip(b'\r\n')
                if file_bytes.endswith(b'--'):
                    file_bytes = file_bytes[:-2]
                while file_bytes.endswith(b'\r\n') or file_bytes.endswith(b'\n'):
                    file_bytes = file_bytes.rstrip(b'\r\n')
                return file_bytes
    
    return None

def lambda_handler(event, context):
    """Handle API Gateway requests. PHI-safe - no logging of file contents."""
    
    # Get request method and path
    http_method = event.get('requestContext', {}).get('http', {}).get('method', 'POST')
    path = event.get('rawPath', event.get('path', '/'))
    
    # Handle CORS preflight
    if http_method == 'OPTIONS':
        return {
            'statusCode': 200,
            'headers': {
                'Access-Control-Allow-Origin': ALLOWED_ORIGIN,
                'Access-Control-Allow-Headers': 'Content-Type',
                'Access-Control-Allow-Methods': 'POST, OPTIONS, GET',
                'Access-Control-Expose-Headers': 'X-Stats-Total, X-Stats-Charge, X-Stats-Deny, X-Stats-Review, X-Stats-Patients, Content-Disposition'
            },
            'body': ''
        }
    
    # Handle health check
    if path.endswith('/health') or path == '/api/health':
        return {
            'statusCode': 200,
            'headers': {
                'Content-Type': 'application/json',
                'Access-Control-Allow-Origin': ALLOWED_ORIGIN
            },
            'body': json.dumps({
                'status': 'healthy',
                'service': 'ChargeScrub',
                'version': '1.0.0'
            })
        }
    
    # Handle file upload
    try:
        # Parse request body
        if event.get('isBase64Encoded'):
            body = base64.b64decode(event['body'])
        else:
            body = event['body'].encode() if isinstance(event['body'], str) else event['body']
        
        # Get content type - handle case-insensitive headers
        headers = event.get('headers', {})
        content_type = headers.get('content-type', '') or headers.get('Content-Type', '')
        
        file_bytes = None
        
        if 'multipart/form-data' in content_type:
            # Extract boundary from content type
            for part in content_type.split(';'):
                part = part.strip()
                if part.startswith('boundary='):
                    boundary = part[9:].strip('"')
                    file_bytes = parse_multipart(body, boundary)
                    break
            
            if not file_bytes:
                return {
                    'statusCode': 400,
                    'headers': {'Content-Type': 'application/json', 'Access-Control-Allow-Origin': ALLOWED_ORIGIN},
                    'body': json.dumps({'error': 'No file found in multipart request'})
                }
        else:
            # Assume raw binary
            file_bytes = body
        
        # Validate we have data
        if not file_bytes or len(file_bytes) < 100:
            return {
                'statusCode': 400,
                'headers': {'Content-Type': 'application/json', 'Access-Control-Allow-Origin': ALLOWED_ORIGIN},
                'body': json.dumps({'error': 'No valid file data received', 'size': len(file_bytes) if file_bytes else 0})
            }
        
        # Process the file
        processed_bytes, stats = process_excel(file_bytes)
        
        # Return processed file
        return {
            'statusCode': 200,
            'headers': {
                'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                'Content-Disposition': f'attachment; filename="ChargeScrub_Output_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"',
                'Access-Control-Allow-Origin': ALLOWED_ORIGIN,
                'Access-Control-Allow-Headers': 'Content-Type',
                'Access-Control-Allow-Methods': 'POST, OPTIONS, GET',
                'X-Stats-Total': str(stats['total']),
                'X-Stats-Charge': str(stats['charge']),
                'X-Stats-Deny': str(stats['deny']),
                'X-Stats-Review': str(stats['review']),
                'X-Stats-Patients': str(stats['patients']),
                'Access-Control-Expose-Headers': 'X-Stats-Total, X-Stats-Charge, X-Stats-Deny, X-Stats-Review, X-Stats-Patients, Content-Disposition'
            },
            'isBase64Encoded': True,
            'body': base64.b64encode(processed_bytes).decode('utf-8')
        }
        
    except Exception as e:
        # Log error type only, not content (PHI safety)
        # Do NOT include str(e) as it may contain patient data
        error_type = type(e).__name__
        return {
            'statusCode': 500,
            'headers': {'Content-Type': 'application/json', 'Access-Control-Allow-Origin': ALLOWED_ORIGIN},
            'body': json.dumps({'error': f'Processing error: {error_type}. Please verify file format.'})
        }
